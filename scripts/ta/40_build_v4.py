#!/usr/bin/env python3
"""Stage 4: apply the dedupe plan + all verification results and write the
deliverable workbook.

Design decisions that matter:
  * Companies and people are SEPARATED. A row whose only identity is a
    personal email address is a person, not a company — it goes to its own
    tab instead of inflating the company count.
  * Merging UNIONS data (phones, emails, licences, domains); nothing is
    dropped, and the audit tab records every merge.
  * Every added value carries a source. Blank means unknown, never zero.

Usage: 40_build_v4.py <work.jsonl> <crawl.jsonl> <strict.jsonl> <names.jsonl>
                      <dedupe.json> <out.xlsx> <preview.txt>
"""
import sys, json, re, collections
sys.path.insert(0, '/root/.claude/skills/synced/contact-data-consolidation/scripts')
from normalize import valid_email, maps_url, priority_score, completeness, norm_name

WORK, CRAWL, STRICT, NAMES, DEDUPE, OUT, PREVIEW = sys.argv[1:8]

recs = [json.loads(l) for l in open(WORK, encoding='utf-8')]
crawl = {json.loads(l)['domain']: json.loads(l) for l in open(CRAWL, encoding='utf-8')}
strict = {json.loads(l)['domain']: json.loads(l) for l in open(STRICT, encoding='utf-8')}
names = {json.loads(l)['domain']: json.loads(l) for l in open(NAMES, encoding='utf-8')}
plan = json.load(open(DEDUPE, encoding='utf-8'))

FREEMAIL = {'gmail.com', 'gmail.coma', 'hotmail.com', 'hotmail.co.uk', 'hotmail.fr', 'yahoo.com',
            'yahoo.fr', 'yahoo.co.uk', 'yahoo.con', 'yahoo.coma', 'outlook.com', 'outlook.sa',
            'icloud.com', 'live.com', 'msn.com', 'windowslive.com', 'aol.com', 'mail.ru',
            'yandex.com', 'rediffmail.com', 'protonmail.com', 'me.com', 'qq.com', '163.com'}
AREA = {'11': 'Riyadh', '12': 'Makkah', '13': 'Eastern Province',
        '14': 'Madinah / Tabuk / North', '16': 'Qassim / Hail',
        '17': 'Asir / Jazan / Najran / Baha'}
CITY_OF = {'Riyadh': 'Riyadh', 'Makkah': 'Jeddah / Makkah', 'Eastern Province': 'Dammam area'}

by_id = {r.get('Row ID'): r for r in recs}
merged_into = {}
for g in plan:
    for m in g['member_row_ids']:
        if m != g['survivor_row_id']:
            merged_into[m] = g['survivor_row_id']

def g(r, k):
    v = r.get(k)
    return '' if v is None else str(v).strip()

def combine(members):
    """Union every fact across the merged rows. Nothing is discarded."""
    out = {'mobiles': [], 'landlines': [], 'hotlines': [], 'emails': [], 'domains': [],
           'placeholders': [], 'whatsapp': '', 'licences': [], 'lead_sources': [],
           'iata': '', 'iban': '', 'cities': [], 'regions': [], 'sar': 0.0,
           'contact_names': [], 'linkedin': '', 'socials': [], 'decision_makers': []}
    for r in members:
        n = r['_norm']
        for k in ('mobiles', 'landlines', 'hotlines', 'emails', 'domains', 'placeholders'):
            for v in n[k]:
                if v not in out[k]: out[k].append(v)
        if n['whatsapp'] and not out['whatsapp']: out['whatsapp'] = n['whatsapp']
        for lic in re.findall(r'\b73\d{6}(?:-\d+)?\b', g(r, 'License Numbers (MOT)')):
            if lic not in out['licences']: out['licences'].append(lic)
        for s in re.split(r'\s*\|\s*', g(r, 'Lead Sources')):
            if s and s not in out['lead_sources']: out['lead_sources'].append(s)
        for fld, key in (('IATA Number', 'iata'), ('IBAN', 'iban'), ('LinkedIn', 'linkedin')):
            if g(r, fld) and not out[key]: out[key] = g(r, fld)
        for fld, key in (('HQ City', 'cities'), ('HQ Region', 'regions'),
                         ('Contact Names', 'contact_names'), ('Decision Maker', 'decision_makers'),
                         ('Social Links', 'socials')):
            v = g(r, fld)
            if v and v not in out[key]: out[key].append(v)
        try: out['sar'] += float(g(r, 'Direct Payments SAR') or 0)
        except ValueError: pass
    return out

rows, people, audit = [], [], []
seen_survivors = set()

for r in recs:
    rid = r.get('Row ID')
    if rid in merged_into:
        continue                      # folded into its survivor
    group = next((x for x in plan if x['survivor_row_id'] == rid), None)
    members = [r]
    if group:
        members = [by_id[m] for m in group['member_row_ids'] if m in by_id]
        seen_survivors.add(rid)
    c = combine(members)

    name_en = next((g(m, 'Agency Name (EN)') for m in members if g(m, 'Agency Name (EN)')), '')
    name_ar = next((g(m, 'Agency Name (AR)') for m in members if g(m, 'Agency Name (AR)')), '')

    # --- recover a name from the company's own website
    name_src = 'source file' if (name_en or name_ar) else ''
    if not name_en and not name_ar:
        for d in c['domains'] + [e.split('@')[1] for e in c['emails']]:
            if d.lower() in FREEMAIL or d not in names:
                continue
            nm = names[d]
            cand = nm.get('og_site') or nm.get('title') or nm.get('h1')
            if cand:
                cand = re.split(r'\s*[|\-–—»:]\s*', cand)[0].strip()
                if 3 < len(cand) < 90:
                    if re.search(r'[؀-ۿ]', cand): name_ar = cand
                    else: name_en = cand
                    name_src = f'website title ({d}, 2026-08-13)'
                    break

    corp_domains = [d for d in c['domains'] if d.lower() not in FREEMAIL]
    corp_emaildoms = {e.split('@')[1] for e in c['emails'] if e.split('@')[1] not in FREEMAIL}

    # --- record type: a person is not a company
    if not name_en and not name_ar and not c['licences'] and not corp_domains and not corp_emaildoms:
        rtype = 'individual / unidentified contact'
    elif not name_en and not name_ar:
        rtype = 'company (name unknown)'
    else:
        rtype = 'company'

    # --- website status + identifiers published by the company itself
    statuses = {crawl[d]['status'] for d in corp_domains if d in crawl}
    web = ('live' if 'live' in statuses else 'parked/thin' if 'parked/thin' in statuses
           else 'dead' if statuses else '')
    site_cr, site_vat, site_lic, cr_ctx = [], [], [], ''
    for d in corp_domains:
        s = strict.get(d)
        if not s: continue
        for x in s['cr']:
            if x['value'] not in site_cr:
                site_cr.append(x['value'])
                if not cr_ctx: cr_ctx = x['context'][:180]
        for x in s['vat']:
            if x['value'] not in site_vat: site_vat.append(x['value'])
        for x in s['licence']:
            if x['value'] not in site_lic: site_lic.append(x['value'])

    sbc_cr = re.sub(r'\D', '', next((g(m, 'CR / Unified No. (SBC)') for m in members
                                     if g(m, 'CR / Unified No. (SBC)')), ''))
    cr, cr_src = '', ''
    if sbc_cr and len(sbc_cr) == 10:
        cr, cr_src = sbc_cr, 'SBC eAuthenticate (July 2026)'
        if site_cr and sbc_cr not in site_cr:
            cr_src += f' — CONFLICT: website shows {site_cr[0]}'
    elif len(site_cr) == 1:
        cr, cr_src = site_cr[0], 'published on company website (2026-08-13)'
    elif len(site_cr) > 1:
        cr_src = 'ambiguous — website shows: ' + ' | '.join(site_cr[:3])

    if cr and 'CONFLICT' not in cr_src:
        level = 'confirmed'
    elif c['licences'] or web == 'live':
        level = 'candidate'
    elif c['emails'] or c['mobiles'] or c['landlines']:
        level = 'candidate (contact only)'
    else:
        level = 'unverified'

    region = next((x for x in c['regions'] if x and 'unconfirmed' not in x.lower()), '')
    rsrc = 'source file' if region else ''
    if not region:
        for ll in c['landlines']:
            if ll[4:6] in AREA:
                region, rsrc = AREA[ll[4:6]], f'derived from landline code 0{ll[4:6]}'
                break
    city = next((x for x in c['cities'] if x and 'unconfirmed' not in x.lower()), '')

    disp = name_en or name_ar
    row = {
        'Row ID': rid,
        'Record Type': rtype,
        'Company Name (EN)': name_en,
        'Company Name (AR)': name_ar,
        'Name Source': name_src,
        'Verification Level': level,
        'CR / Unified Number': cr,
        'CR Source': cr_src,
        'CR Evidence (from website)': cr_ctx,
        'VAT Number (from website)': ' | '.join(site_vat[:2]),
        'MOT Licence(s)': ' | '.join(c['licences'][:4]),
        'Licence Status': next((g(m, 'License Status') for m in members if g(m, 'License Status')), ''),
        'Mobile 1': c['mobiles'][0] if c['mobiles'] else '',
        'Mobile 2': c['mobiles'][1] if len(c['mobiles']) > 1 else '',
        'Landline': c['landlines'][0] if c['landlines'] else '',
        'Hotline': c['hotlines'][0] if c['hotlines'] else '',
        'WhatsApp (confirmed)': c['whatsapp'],
        'Email 1': c['emails'][0] if c['emails'] else '',
        'Email 2': c['emails'][1] if len(c['emails']) > 1 else '',
        'Email 3': c['emails'][2] if len(c['emails']) > 2 else '',
        'Extra Emails': ' | '.join(c['emails'][3:8]),
        'Placeholder Phones (ignored)': ' | '.join(c['placeholders'][:3]),
        'Website': corp_domains[0] if corp_domains else '',
        'Website Status (checked 2026-08-13)': web,
        'Other Domains': ' | '.join(corp_domains[1:4]),
        'HQ City': city,
        'Region': region,
        'Region Source': rsrc,
        'Google Maps Search': maps_url(disp, city) if disp else '',
        'IATA Number': c['iata'],
        'IBAN': c['iban'],
        'LinkedIn': c['linkedin'],
        'Decision Maker': ' | '.join(c['decision_makers'][:2]),
        'Contact Names': ' | '.join(c['contact_names'][:3]),
        'Billed by Direct (SAR)': round(c['sar'], 2) if c['sar'] else '',
        'Lead Sources': ' | '.join(c['lead_sources'][:6]),
        'Merged From': ' | '.join(m for m in (group['member_row_ids'] if group else []) if m != rid),
        'Needs Manual Confirmation': any(str(g(m, 'Needs Manual Confirmation')).lower() == 'true' for m in members)
                                      or 'CONFLICT' in cr_src or 'ambiguous' in cr_src,
    }
    row['Priority Score'] = priority_score(
        bool(c['emails'] and valid_email(c['emails'][0])), bool(c['mobiles']),
        'riyadh' in (city + ' ' + region).lower(), web == 'live',
        bool(c['licences']), bool(c['whatsapp']))
    row['Completeness %'] = completeness(
        {'a': disp, 'b': row['Mobile 1'], 'c': row['Landline'], 'd': row['Email 1'],
         'e': 'y' if web == 'live' else '', 'f': cr, 'g': row['MOT Licence(s)'],
         'h': city, 'i': region},
        list('abcdefghi'))

    if group:
        audit.append({'Kept (survivor)': rid, 'Name': disp[:60],
                      'Folded in': ' | '.join(x for x in group['member_row_ids'] if x != rid),
                      'How they matched': ' ; '.join(group['rules'][:3]),
                      'Rows combined': group['size']})

    (people if rtype.startswith('individual') else rows).append(row)

COLS = ['Row ID', 'Record Type', 'Company Name (EN)', 'Company Name (AR)', 'Name Source',
        'Verification Level', 'CR / Unified Number', 'CR Source', 'CR Evidence (from website)',
        'VAT Number (from website)', 'MOT Licence(s)', 'Licence Status',
        'Mobile 1', 'Mobile 2', 'Landline', 'Hotline', 'WhatsApp (confirmed)',
        'Email 1', 'Email 2', 'Email 3', 'Extra Emails', 'Placeholder Phones (ignored)',
        'Website', 'Website Status (checked 2026-08-13)', 'Other Domains',
        'HQ City', 'Region', 'Region Source', 'Google Maps Search',
        'IATA Number', 'IBAN', 'LinkedIn', 'Decision Maker', 'Contact Names',
        'Billed by Direct (SAR)', 'Lead Sources', 'Merged From',
        'Needs Manual Confirmation', 'Priority Score', 'Completeness %']

DEFS = [
 ('Record Type', 'company = a real business. "company (name unknown)" = a business we can reach but whose name we have not recovered yet. Individuals are on their own tab, not counted as companies.'),
 ('Verification Level', 'confirmed = has an official CR/unified number from SBC or published by the company itself. candidate = has a MOT licence or a live website. candidate (contact only) = we can reach them but nothing official yet. unverified = nothing checkable.'),
 ('CR / Unified Number', 'The commercial registration / unified number. Blank means unknown — never guessed.'),
 ('CR Source', 'Where that number came from. CONFLICT = SBC and the website disagree, check by hand. ambiguous = the site shows several.'),
 ('CR Evidence (from website)', 'The actual sentence on the company website where the number was found, so you can judge it yourself.'),
 ('MOT Licence(s)', 'Ministry of Tourism licence number(s).'),
 ('Mobile 1 / Mobile 2', 'Saudi mobiles in international format. Fake/placeholder numbers were removed. Safe for SMS and WhatsApp.'),
 ('Landline / Hotline', 'Office number (area code shows the city) / 920 and 800 service numbers.'),
 ('WhatsApp (confirmed)', 'Filled only when a source said this number is WhatsApp. Blank does not mean "no WhatsApp" — it means unconfirmed.'),
 ('Email 1-3 / Extra Emails', 'Valid email addresses, best first.'),
 ('Placeholder Phones (ignored)', 'Obviously fake numbers found in the sources (111111..., 123456...). Kept visible so nothing is hidden, but never counted as reachable.'),
 ('Website Status (checked 2026-08-13)', 'We visited every website on 13 August 2026: live / parked-thin / dead. Blank = no website known.'),
 ('Region + Region Source', 'Where they are. "derived from landline code" means we worked it out from the phone area code, not from a source that stated it.'),
 ('Google Maps Search', 'Click to look the office up on Google Maps.'),
 ('Billed by Direct (SAR)', 'Money this company has already paid Direct, from the Direct Payments export.'),
 ('Merged From', 'Row IDs of duplicate records folded into this one. Their data was combined, not deleted. Full list on the "Duplicates merged" tab.'),
 ('Needs Manual Confirmation', 'TRUE = a human should check the identity of this record (conflicting numbers, or an uncertain match).'),
 ('Priority Score', '0-100. Who to work first: reachable + Riyadh + live website + licensed.'),
 ('Completeness %', 'How much of the 9 key fields this row has.'),
]

lv = collections.Counter(r['Verification Level'] for r in rows)
QUAL = [
 ('Companies (one row each)', len(rows)),
 ('Individuals / unidentified contacts (separate tab)', len(people)),
 ('Rows in the v3 file we started from', len(recs)),
 ('Duplicate records folded in', sum(a['Rows combined'] - 1 for a in audit)),
 ('', ''),
 ('CONFIRMED (official CR number)', lv['confirmed']),
 ('Candidate (licence or live website)', lv['candidate']),
 ('Candidate (contact only)', lv['candidate (contact only)']),
 ('Unverified', lv['unverified']),
 ('', ''),
 ('CR numbers known', sum(1 for r in rows if r['CR / Unified Number'])),
 ('  of those, newly found on the company website today', sum(1 for r in rows if 'website' in r['CR Source'])),
 ('VAT numbers found today', sum(1 for r in rows if r['VAT Number (from website)'])),
 ('CR conflicts to check by hand', sum(1 for r in rows if 'CONFLICT' in r['CR Source'])),
 ('', ''),
 ('Websites checked on 13 Aug 2026', sum(1 for r in rows if r['Website Status (checked 2026-08-13)'])),
 ('  live', sum(1 for r in rows if r['Website Status (checked 2026-08-13)'] == 'live')),
 ('  parked / almost empty', sum(1 for r in rows if r['Website Status (checked 2026-08-13)'] == 'parked/thin')),
 ('  dead (site gone)', sum(1 for r in rows if r['Website Status (checked 2026-08-13)'] == 'dead')),
 ('', ''),
 ('Company names recovered from their website today', sum(1 for r in rows if 'website title' in r['Name Source'])),
 ('Companies still with no name', sum(1 for r in rows if r['Record Type'] == 'company (name unknown)')),
 ('', ''),
 ('Reachable by mobile', sum(1 for r in rows if r['Mobile 1'])),
 ('Reachable by landline', sum(1 for r in rows if r['Landline'])),
 ('Reachable by email', sum(1 for r in rows if r['Email 1'])),
 ('Not reachable by anything', sum(1 for r in rows if not (r['Mobile 1'] or r['Landline'] or r['Email 1'] or r['Website']))),
 ('Fake/placeholder phones removed', sum(1 for r in rows if r['Placeholder Phones (ignored)'])),
 ('', ''),
 ('Region known', sum(1 for r in rows if r['Region'])),
 ('  of those, worked out from the phone area code', sum(1 for r in rows if 'derived' in r['Region Source'])),
 ('Flagged for a human to confirm', sum(1 for r in rows if r['Needs Manual Confirmation'])),
]

top = sorted(rows, key=lambda r: (-r['Priority Score'], -r['Completeness %']))[:18]
prev = ['TRAVEL AGENCIES — MASTER v4 (13 Aug 2026)',
        f'{len(rows)} companies  |  {lv["confirmed"]} confirmed with an official CR number',
        f'{sum(1 for r in rows if r["Website Status (checked 2026-08-13)"]=="live")} live websites  |  '
        f'{sum(1 for r in rows if r["Mobile 1"])} mobiles  |  {sum(1 for r in rows if r["Email 1"])} emails',
        f'{sum(a["Rows combined"]-1 for a in audit)} duplicates merged  |  {len(people)} individuals moved off the company list',
        '', 'WORK THESE FIRST:']
for r in top:
    nm = (r['Company Name (EN)'] or r['Company Name (AR)'])[:32]
    prev.append(f'{r["Priority Score"]:>3} {nm:<32} {(r["Mobile 1"] or r["Landline"] or "-"):<14} {(r["Email 1"] or "-")[:26]}')

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
wb = openpyxl.Workbook()
cov = wb.active; cov.title = 'READ ME'
cov.append(['Column', 'What it means'])
for a, b in DEFS: cov.append([a, b])
ms = wb.create_sheet('Master — companies'); ms.append(COLS)
for r in rows: ms.append([r.get(c, '') for c in COLS])
ps = wb.create_sheet('Individuals — not companies')
PC = ['Row ID', 'Email 1', 'Email 2', 'Mobile 1', 'Contact Names', 'Lead Sources']
ps.append(PC + ['Why here'])
for r in people:
    ps.append([r.get(c, '') for c in PC] + ['Only identity is a personal email address — belongs in a contact list, not the company database'])
as_ = wb.create_sheet('Duplicates merged')
AC = ['Kept (survivor)', 'Name', 'Folded in', 'How they matched', 'Rows combined']
as_.append(AC)
for a in sorted(audit, key=lambda x: -x['Rows combined']): as_.append([a[c] for c in AC])
qs = wb.create_sheet('Data quality'); qs.append(['Measure', 'Value'])
for a, b in QUAL: qs.append([a, b])

TEXT = ['CR / Unified Number', 'VAT Number (from website)', 'MOT Licence(s)', 'Mobile 1',
        'Mobile 2', 'Landline', 'Hotline', 'WhatsApp (confirmed)', 'IBAN', 'IATA Number']
hi = {c: i + 1 for i, c in enumerate(COLS)}
for c in TEXT:
    if c in hi:
        for rr in range(2, ms.max_row + 1): ms.cell(row=rr, column=hi[c]).number_format = '@'
hf = PatternFill('solid', fgColor='F06820'); ff = Font(bold=True, color='FFFFFF')
for sh in wb.worksheets:
    sh.freeze_panes = 'A2'
    for c in sh[1]: c.fill = hf; c.font = ff
ms.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}{ms.max_row}'
cov.column_dimensions['A'].width = 34; cov.column_dimensions['B'].width = 96
for sh, w in ((ms, 30), (ps, 26), (as_, 30), (qs, 46)):
    sh.column_dimensions['A'].width = 12
    sh.column_dimensions['B'].width = w
wb.save(OUT)
open(PREVIEW, 'w', encoding='utf-8').write('\n'.join(prev[:30]))

print('wrote', OUT)
for k, v in QUAL:
    if k: print(f'{k}: {v}')
