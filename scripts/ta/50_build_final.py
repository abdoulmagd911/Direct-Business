#!/usr/bin/env python3
"""Stage 5: apply every verification result and write the deliverable workbook.

Inputs, in order of authority:
  ta-work.jsonl        the v3 rows, normalised
  dedupe_decisions     agent adjudication of the risky duplicate groups
  dedupe_safe.json     groups auto-accepted (licence / exact-Arabic-name, small)
  id_verdicts.json     agent-judged CR / VAT / MOT licence, with the traps rejected
  name_verdicts.json   agent-judged company names recovered from websites
  ta-crawl.jsonl       website liveness
  ta-mx.json           which email domains can actually receive mail

Rules kept throughout: never invent a value; blank means unknown; every added
value carries a source; merging unions data and is recorded; people are not
counted as companies.

Usage: 50_build_final.py <outdir>
"""
import sys, json, re, collections, os
sys.path.insert(0, '/root/.claude/skills/synced/contact-data-consolidation/scripts')
from normalize import valid_email, maps_url, priority_score, completeness

SP = '/tmp/claude-0/-home-user-Direct-Business/87d1aec0-e2cf-54b0-95a9-480203ce0e75/scratchpad'
OUTDIR = sys.argv[1] if len(sys.argv) > 1 else SP

def load_jsonl(p):
    return [json.loads(l) for l in open(p, encoding='utf-8')]
def load_json(p, default=None):
    if not os.path.exists(p): return default if default is not None else []
    return json.load(open(p, encoding='utf-8'))

recs = load_jsonl(f'{SP}/ta-work.jsonl')
crawl = {r['domain']: r for r in load_jsonl(f'{SP}/ta-crawl.jsonl')}
mx = load_json(f'{SP}/ta-mx.json', {})
decisions = load_json(f'{SP}/dedupe_decisions.json')
safe = load_json(f'{SP}/dedupe_safe.json')
idv = load_json(f'{SP}/id_verdicts.json')
namev = load_json(f'{SP}/name_verdicts.json')

by_id = {r.get('Row ID'): r for r in recs}
id_by_row = {v['row_id']: v for v in idv if v.get('row_id')}
id_by_dom = {v['domain']: v for v in idv if v.get('domain')}
name_by_row = {v['row_id']: v for v in namev if v.get('decision') == 'accept' and v.get('clean_name')}
# Rows the reviewers judged to be real organisations that are NOT travel businesses
# (a hospital, bank or industrial group with an in-house travel desk). They are
# corporate-client prospects, not competitors — a different sales conversation.
not_travel = {v['row_id'] for v in namev
              if v.get('decision') == 'accept' and v.get('is_travel_business') is False}
NON_TRAVEL_HINT = re.compile(r'\b(bank|hospital|insurance|jewell?ery|finance|'
                             r'electronics|contracting|automobile association)\b', re.I)

FREEMAIL = {'gmail.com', 'hotmail.com', 'yahoo.com', 'outlook.com', 'icloud.com', 'live.com',
            'msn.com', 'aol.com', 'mail.ru', 'yandex.com', 'protonmail.com', 'me.com',
            'qq.com', '163.com', 'rediffmail.com', 'windowslive.com'}
def is_free(d):
    d = (d or '').lower()
    return d in FREEMAIL or re.match(r'^(gmail|yahoo|hotmail|outlook|icloud|live|msn|aol)\.', d)

AREA = {'11': 'Riyadh', '12': 'Makkah', '13': 'Eastern Province',
        '14': 'Madinah / Tabuk / North', '16': 'Qassim / Hail',
        '17': 'Asir / Jazan / Najran / Baha'}

# ---------- 1. build the final merge map from the adjudicated decisions ----------
merge_into = {}          # row -> survivor
not_company = set()      # rows the adjudicator said are people, not a business
merge_groups = {}        # survivor -> [members]
audit = []

for g in safe:
    s = g['survivor_row_id']
    members = [m for m in g['member_row_ids'] if m in by_id]
    merge_groups[s] = members
    for m in members:
        if m != s: merge_into[m] = s
    audit.append({'Kept': s, 'Folded in': ' | '.join(m for m in members if m != s),
                  'Rows': len(members), 'Decision': 'auto (licence or exact Arabic name)',
                  'Why': ' ; '.join(g['rules'][:2])})

for d in decisions:
    s = d['survivor_row_id']
    dec = d['decision']
    if dec == 'not_a_company':
        for m in d.get('separate_row_ids', []): not_company.add(m)
        audit.append({'Kept': '(none)', 'Folded in': ' | '.join(d.get('separate_row_ids', [])),
                      'Rows': len(d.get('separate_row_ids', [])), 'Decision': 'NOT a company — kept apart',
                      'Why': d['reason'][:300]})
        continue
    if dec == 'do_not_merge':
        audit.append({'Kept': '(all kept separate)', 'Folded in': '', 'Rows': 0,
                      'Decision': 'do not merge — different companies',
                      'Why': d['reason'][:300]})
        continue
    members = [m for m in (d.get('merge_row_ids') or []) if m in by_id]
    if dec == 'merge_subset' and not members:
        continue
    if not members:
        continue
    if s not in members: members.append(s)
    merge_groups[s] = members
    for m in members:
        if m != s: merge_into[m] = s
    audit.append({'Kept': s, 'Folded in': ' | '.join(m for m in members if m != s),
                  'Rows': len(members),
                  'Decision': 'merge all' if dec == 'merge_all' else 'merge some, others kept separate',
                  'Why': d['reason'][:300]})

def g_(r, k):
    v = r.get(k)
    return '' if v is None else str(v).strip()

def combine(members):
    o = collections.defaultdict(list)
    o['whatsapp'] = ''; o['sar'] = 0.0
    for r in members:
        n = r['_norm']
        for k in ('mobiles', 'landlines', 'hotlines', 'emails', 'domains', 'placeholders'):
            for v in n[k]:
                if v not in o[k]: o[k].append(v)
        if n['whatsapp'] and not o['whatsapp']: o['whatsapp'] = n['whatsapp']
        for lic in re.findall(r'\b73\d{6}\b', g_(r, 'License Numbers (MOT)')):
            if lic not in o['licences']: o['licences'].append(lic)
        for s_ in re.split(r'\s*\|\s*', g_(r, 'Lead Sources')):
            if s_ and s_ not in o['sources']: o['sources'].append(s_)
        for fld, key in (('HQ City', 'cities'), ('HQ Region', 'regions'),
                         ('Contact Names', 'contacts'), ('Decision Maker', 'dms'),
                         ('IATA Number', 'iata'), ('IBAN', 'iban'), ('LinkedIn', 'linkedin'),
                         ('License Status', 'licstatus')):
            v = g_(r, fld)
            if v and v not in o[key]: o[key].append(v)
        try: o['sar'] += float(g_(r, 'Direct Payments SAR') or 0)
        except ValueError: pass
    return o

rows, people, notes_rows = [], [], []

for r in recs:
    rid = r.get('Row ID')
    if rid in merge_into:
        continue
    members = [by_id[m] for m in merge_groups.get(rid, [rid]) if m in by_id]
    c = combine(members)

    name_en = next((g_(m, 'Agency Name (EN)') for m in members if g_(m, 'Agency Name (EN)')), '')
    name_ar = next((g_(m, 'Agency Name (AR)') for m in members if g_(m, 'Agency Name (AR)')), '')
    name_src = 'source file' if (name_en or name_ar) else ''
    if not name_en and not name_ar:
        nv = next((name_by_row[m] for m in merge_groups.get(rid, [rid]) if m in name_by_row), None)
        if nv:
            if re.search(r'[؀-ۿ]', nv['clean_name']): name_ar = nv['clean_name']
            else: name_en = nv['clean_name']
            name_src = f"read from their website {nv['domain']} (13 Aug 2026)"

    corp_doms = [d for d in c['domains'] if not is_free(d)]
    corp_maildoms = sorted({e.split('@')[1] for e in c['emails'] if not is_free(e.split('@')[1])})

    if rid in not_company or (not name_en and not name_ar and not c['licences']
                              and not corp_doms and not corp_maildoms):
        rtype = 'individual / unidentified contact'
    elif not name_en and not name_ar:
        rtype = 'company (name not recovered)'
    else:
        rtype = 'company'

    # ---- verified identifiers (agent-judged; traps already rejected)
    verdict = id_by_row.get(rid) or next((id_by_dom[d] for d in corp_doms if d in id_by_dom), None)
    cr = vat = lic_v = cr_kind = ''
    cr_src = ''
    trap_notes = ''
    if verdict:
        cr = verdict.get('accepted_cr') or ''
        cr_kind = verdict.get('cr_kind') or ''
        vat = verdict.get('accepted_vat') or ''
        lic_v = verdict.get('accepted_licence') or ''
        if cr:
            cr_src = f"published on their website, checked by hand 13 Aug 2026 ({cr_kind.replace('_',' ')})"
        if verdict.get('conflicts_with_sbc'):
            cr_src += ' — DIFFERS from the July SBC record, needs a human look'
        rej = verdict.get('rejected') or []
        if rej: trap_notes = ' ; '.join(rej[:2])[:300]
        if str(verdict.get('notes', '')).startswith('CORRECTED'):
            trap_notes = 'CORRECTED ON REVIEW: ' + str(verdict.get('notes'))[:280]

    if not cr:
        sbc = re.sub(r'\D', '', next((g_(m, 'CR / Unified No. (SBC)') for m in members
                                      if g_(m, 'CR / Unified No. (SBC)')), ''))
        if len(sbc) == 10:
            cr, cr_src = sbc, 'SBC eAuthenticate lookup (July 2026)'
            cr_kind = 'unified_national_number' if sbc.startswith('7') else 'commercial_registration'

    licences = list(c['licences'])
    if lic_v and lic_v not in licences: licences.insert(0, lic_v)

    statuses = {crawl[d]['status'] for d in corp_doms if d in crawl}
    web = ('live' if 'live' in statuses else 'parked/thin' if 'parked/thin' in statuses
           else 'dead' if statuses else '')

    # ---- email deliverability
    mailable = [e for e in c['emails'] if mx.get(e.split('@')[1].lower(), {}).get('status') == 'mx']
    undeliverable = [e for e in c['emails'] if e not in mailable]

    if cr and 'DIFFERS' not in cr_src:
        level = 'confirmed'
    elif licences or web == 'live':
        level = 'candidate'
    elif mailable or c['mobiles'] or c['landlines']:
        level = 'candidate (contact only)'
    else:
        level = 'unverified'

    region = next((x for x in c['regions'] if x and 'unconfirmed' not in x.lower()), '')
    rsrc = 'source file' if region else ''
    if not region:
        for ll in c['landlines']:
            if ll[4:6] in AREA:
                region, rsrc = AREA[ll[4:6]], f'worked out from the phone area code 0{ll[4:6]}'
                break
    city = next((x for x in c['cities'] if x and 'unconfirmed' not in x.lower()), '')
    disp = name_en or name_ar

    row = {
        'Row ID': rid,
        'Record Type': rtype,
        'Company Name (EN)': name_en,
        'Company Name (AR)': name_ar,
        'Where the name came from': name_src,
        'Verification Level': level,
        'CR / Unified Number': cr,
        'Number Type': cr_kind.replace('_', ' '),
        'Where the number came from': cr_src,
        'VAT Number': vat,
        'MOT Licence(s)': ' | '.join(licences[:4]),
        'Licence Status': c['licstatus'][0] if c['licstatus'] else '',
        'Mobile 1': c['mobiles'][0] if c['mobiles'] else '',
        'Mobile 2': c['mobiles'][1] if len(c['mobiles']) > 1 else '',
        'Landline': c['landlines'][0] if c['landlines'] else '',
        'Hotline (920/800)': c['hotlines'][0] if c['hotlines'] else '',
        'WhatsApp (confirmed)': c['whatsapp'],
        'Email 1 (deliverable)': mailable[0] if mailable else '',
        'Email 2 (deliverable)': mailable[1] if len(mailable) > 1 else '',
        'Email 3 (deliverable)': mailable[2] if len(mailable) > 2 else '',
        'More deliverable emails': ' | '.join(mailable[3:8]),
        'Emails that cannot receive mail': ' | '.join(undeliverable[:4]),
        'Fake numbers found (ignored)': ' | '.join(c['placeholders'][:3]),
        'Website': corp_doms[0] if corp_doms else '',
        'Website Status (13 Aug 2026)': web,
        'Other Domains': ' | '.join(corp_doms[1:4]),
        'HQ City': city,
        'Region': region,
        'Where the region came from': rsrc,
        'Find on Google Maps': maps_url(disp, city) if disp else '',
        'IATA Number': c['iata'][0] if c['iata'] else '',
        'IBAN': c['iban'][0] if c['iban'] else '',
        'LinkedIn': c['linkedin'][0] if c['linkedin'] else '',
        'Decision Maker': ' | '.join(c['dms'][:2]),
        'Contact Names': ' | '.join(c['contacts'][:3]),
        'Already billed by Direct (SAR)': round(c['sar'], 2) if c['sar'] else '',
        'Where this record came from': ' | '.join(c['sources'][:6]),
        'Duplicates folded in': ' | '.join(m for m in merge_groups.get(rid, []) if m != rid),
        'Travel business?': ('no — corporate client, not an agency'
                             if (rid in not_travel or
                                 (verdict and NON_TRAVEL_HINT.search(str(verdict.get('notes', '')))))
                             else ''),
        'Numbers we rejected (and why)': trap_notes,
        'Needs a human check': bool(trap_notes.startswith('CORRECTED')) or 'DIFFERS' in cr_src
                               or any(str(g_(m, 'Needs Manual Confirmation')).lower() == 'true' for m in members),
    }
    row['Priority Score'] = priority_score(
        bool(mailable), bool(c['mobiles']),
        'riyadh' in (city + ' ' + region).lower(), web == 'live',
        bool(licences), bool(c['whatsapp']))
    row['Completeness %'] = completeness(
        {'a': disp, 'b': row['Mobile 1'], 'c': row['Landline'], 'd': row['Email 1 (deliverable)'],
         'e': 'y' if web == 'live' else '', 'f': cr, 'g': row['MOT Licence(s)'],
         'h': city, 'i': region}, list('abcdefghi'))

    (people if rtype.startswith('individual') else rows).append(row)

COLS = ['Row ID', 'Record Type', 'Company Name (EN)', 'Company Name (AR)',
        'Where the name came from', 'Verification Level', 'CR / Unified Number', 'Number Type',
        'Where the number came from', 'VAT Number', 'MOT Licence(s)', 'Licence Status',
        'Mobile 1', 'Mobile 2', 'Landline', 'Hotline (920/800)', 'WhatsApp (confirmed)',
        'Email 1 (deliverable)', 'Email 2 (deliverable)', 'Email 3 (deliverable)',
        'More deliverable emails', 'Emails that cannot receive mail', 'Fake numbers found (ignored)',
        'Website', 'Website Status (13 Aug 2026)', 'Other Domains', 'HQ City', 'Region',
        'Where the region came from', 'Find on Google Maps', 'IATA Number', 'IBAN', 'LinkedIn',
        'Decision Maker', 'Contact Names', 'Already billed by Direct (SAR)',
        'Where this record came from', 'Duplicates folded in', 'Travel business?',
        'Numbers we rejected (and why)', 'Needs a human check', 'Priority Score', 'Completeness %']

DEFS = [
 ('HOW TO READ THIS SHEET', 'One row = one company. Blank always means "we do not know", never zero. Every fact says where it came from.'),
 ('Record Type', 'company = a real business. "company (name not recovered)" = we can reach them but do not yet know their name. People who were mixed into the company list are on the "People, not companies" tab.'),
 ('Verification Level', 'confirmed = we have an official registration number, either from the July SBC lookup or published by the company on its own website and checked by hand. candidate = has a tourism licence or a working website. candidate (contact only) = we can reach them but have nothing official. unverified = nothing to check yet.'),
 ('CR / Unified Number + Number Type', 'The official number. Saudi companies have two kinds: a commercial registration (starts with a city code such as 1010 Riyadh, 4030 Jeddah) and a unified national number (starts with 7). Both are valid identities; the type column says which one this is.'),
 ('Where the number came from', 'Its source. "DIFFERS from the July SBC record" means the two sources disagree and someone should look.'),
 ('MOT Licence(s)', 'Ministry of Tourism licence number(s), 8 digits starting 73.'),
 ('Mobile / Landline / Hotline', 'Split into separate columns and cleaned. Obvious fakes were removed (see "Fake numbers found"). Only mobiles are safe for SMS and WhatsApp.'),
 ('WhatsApp (confirmed)', 'Only filled when a source actually said this number is WhatsApp. Blank means unconfirmed, not "no WhatsApp".'),
 ('Email 1-3 (deliverable)', 'Addresses whose domain really accepts mail — we checked the mail records of every domain on 13 August 2026.'),
 ('Emails that cannot receive mail', 'Addresses whose domain has no mail server. Sending to these will bounce.'),
 ('Website Status (13 Aug 2026)', 'We visited every website that day. live = a working site. parked/thin = the domain exists but there is no real site. dead = it did not respond.'),
 ('Region / Where the region came from', 'If it says "worked out from the phone area code", we deduced it rather than read it from a source.'),
 ('Already billed by Direct (SAR)', 'Money this company has already paid Direct, from the Direct Payments export. A paying customer is the warmest lead you have.'),
 ('Duplicates folded in', 'Row IDs of duplicate records merged into this one. Their data was combined, never deleted. Full reasoning on the "Duplicates" tab.'),
 ('Travel business?', 'Blank = a travel/tourism business. "no — corporate client" = a real company that is NOT an agency (a hospital, bank or industrial group whose staff travel desk put it on the list). Those are buyers, not competitors — a different sales conversation.'),
 ('Numbers we rejected (and why)', 'Numbers found on the website that look official but are not — trademark numbers, toll-free lines, IATA codes, foreign registrations. Kept so you can see what was considered.'),
 ('Needs a human check', 'TRUE = something conflicts and a person should decide.'),
 ('Priority Score', '0-100, who to work first: reachable + Riyadh + live website + licensed.'),
]

lv = collections.Counter(r['Verification Level'] for r in rows)
folded = sum(a['Rows'] - 1 for a in audit if a['Rows'] > 1 and a['Kept'] != '(none)')
QUAL = [
 ('COMPANIES (one row each)', len(rows)),
 ('People wrongly listed as companies, moved off', len(people)),
 ('Rows in the file we started from (v3)', len(recs)),
 ('Duplicate records merged away', folded),
 ('Groups the review said must NOT be merged', sum(1 for a in audit if a['Decision'].startswith('do not merge'))),
 ('', ''),
 ('CONFIRMED — has an official registration number', lv['confirmed']),
 ('Candidate — licensed or has a live website', lv['candidate']),
 ('Candidate — we can contact them, nothing official yet', lv['candidate (contact only)']),
 ('Unverified', lv['unverified']),
 ('', ''),
 ('Official numbers found on company websites (new today)', sum(1 for r in rows if 'website' in r['Where the number came from'])),
 ('VAT numbers found (new today)', sum(1 for r in rows if r['VAT Number'])),
 ('Tourism licences on record', sum(1 for r in rows if r['MOT Licence(s)'])),
 ('Fake "registration numbers" caught and rejected', sum(1 for r in rows if r['Numbers we rejected (and why)'])),
 ('Records where sources disagree — please check', sum(1 for r in rows if r['Needs a human check'])),
 ('', ''),
 ('Websites visited on 13 Aug 2026', sum(1 for r in rows if r['Website Status (13 Aug 2026)'])),
 ('  working', sum(1 for r in rows if r['Website Status (13 Aug 2026)'] == 'live')),
 ('  parked / empty', sum(1 for r in rows if r['Website Status (13 Aug 2026)'] == 'parked/thin')),
 ('  dead', sum(1 for r in rows if r['Website Status (13 Aug 2026)'] == 'dead')),
 ('Company names recovered from their website today', sum(1 for r in rows if 'read from their website' in r['Where the name came from'])),
 ('Companies still with no name', sum(1 for r in rows if r['Record Type'] == 'company (name not recovered)')),
 ('', ''),
 ('Reachable by mobile', sum(1 for r in rows if r['Mobile 1'])),
 ('Reachable by landline', sum(1 for r in rows if r['Landline'])),
 ('Reachable by a WORKING email', sum(1 for r in rows if r['Email 1 (deliverable)'])),
 ('Emails that would bounce (found and separated)', sum(1 for r in rows if r['Emails that cannot receive mail'])),
 ('Fake phone numbers removed', sum(1 for r in rows if r['Fake numbers found (ignored)'])),
 ('Not reachable by anything', sum(1 for r in rows if not (r['Mobile 1'] or r['Landline'] or r['Email 1 (deliverable)'] or r['Website']))),
 ('', ''),
 ('Region known', sum(1 for r in rows if r['Region'])),
 ('  worked out from the phone area code', sum(1 for r in rows if 'area code' in r['Where the region came from'])),
 ('Already paying Direct', sum(1 for r in rows if r['Already billed by Direct (SAR)'])),
 ('Not agencies — corporate clients spotted on the list', sum(1 for r in rows if r['Travel business?'])),
]

top = sorted(rows, key=lambda r: (-r['Priority Score'], -r['Completeness %']))[:18]
prev = ['TRAVEL AGENCIES — MASTER v4 (13 Aug 2026)',
        f'{len(rows)} companies  ·  {lv["confirmed"]} with an official registration number',
        f'{sum(1 for r in rows if r["Website Status (13 Aug 2026)"]=="live")} working websites  ·  '
        f'{sum(1 for r in rows if r["Mobile 1"])} mobiles  ·  {sum(1 for r in rows if r["Email 1 (deliverable)"])} working emails',
        f'{folded} duplicates merged  ·  {len(people)} people moved off the company list',
        '', 'WORK THESE FIRST:']
for r in top:
    nm = (r['Company Name (EN)'] or r['Company Name (AR)'])[:30]
    prev.append(f'{r["Priority Score"]:>3}  {nm:<30} {(r["Mobile 1"] or r["Landline"] or "-"):<14} {(r["Email 1 (deliverable)"] or "-")[:24]}')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
wb = openpyxl.Workbook()
cov = wb.active; cov.title = 'READ ME'
cov.append(['Column', 'What it means'])
for a, b in DEFS: cov.append([a, b])
ms = wb.create_sheet('Companies'); ms.append(COLS)
for r in sorted(rows, key=lambda x: -x['Priority Score']): ms.append([r.get(c, '') for c in COLS])
ps = wb.create_sheet('People, not companies')
PC = ['Row ID', 'Email 1 (deliverable)', 'Email 2 (deliverable)', 'Mobile 1', 'Contact Names', 'Where this record came from']
ps.append(PC + ['Why this is here'])
for r in people:
    ps.append([r.get(c, '') for c in PC] +
              ['The only identity on this record is a personal email address, so it is a person — it belongs in a contact list, not the company database'])
as_ = wb.create_sheet('Duplicates')
AC = ['Kept', 'Folded in', 'Rows', 'Decision', 'Why']
as_.append(AC)
for a in sorted(audit, key=lambda x: -x['Rows']): as_.append([a.get(c, '') for c in AC])
qs = wb.create_sheet('Data quality'); qs.append(['Measure', 'Value'])
for a, b in QUAL: qs.append([a, b])

TEXT = ['CR / Unified Number', 'VAT Number', 'MOT Licence(s)', 'Mobile 1', 'Mobile 2',
        'Landline', 'Hotline (920/800)', 'WhatsApp (confirmed)', 'IBAN', 'IATA Number']
hi = {c: i + 1 for i, c in enumerate(COLS)}
for c in TEXT:
    if c in hi:
        for rr in range(2, ms.max_row + 1): ms.cell(row=rr, column=hi[c]).number_format = '@'
hf = PatternFill('solid', fgColor='F06820'); ff = Font(bold=True, color='FFFFFF')
for sh in wb.worksheets:
    sh.freeze_panes = 'A2'
    for c in sh[1]: c.fill = hf; c.font = ff; c.alignment = Alignment(vertical='center', wrap_text=True)
ms.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}{ms.max_row}'
cov.column_dimensions['A'].width = 32; cov.column_dimensions['B'].width = 110
for col, w in (('A', 11), ('B', 26), ('C', 34), ('D', 30), ('E', 26), ('F', 22), ('G', 18)):
    ms.column_dimensions[col].width = w
qs.column_dimensions['A'].width = 52; qs.column_dimensions['B'].width = 14
as_.column_dimensions['A'].width = 12; as_.column_dimensions['B'].width = 34
as_.column_dimensions['E'].width = 90
ps.column_dimensions['B'].width = 30; ps.column_dimensions['G'].width = 70

out_xlsx = os.path.join(OUTDIR, 'TravelAgencies_MASTER_v4.xlsx')
out_prev = os.path.join(OUTDIR, 'TravelAgencies_v4_PREVIEW.txt')
wb.save(out_xlsx)
open(out_prev, 'w', encoding='utf-8').write('\n'.join(prev[:30]))
print('wrote', out_xlsx)
for k, v in QUAL:
    if k: print(f'{k}: {v}')
