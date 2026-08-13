#!/usr/bin/env python3
"""Stage 6: emit ONE flat table — every record, one header row, no extra tabs.

The workbook version splits companies, people, merge audit and quality onto
separate tabs. The owner works in a single table, so this writes exactly that:
one header, every record beneath it, with `Record Type` telling you what each
row is. Nothing is hidden on another sheet.

Rows folded in as duplicates stay listed on their survivor's row (in
`Duplicates folded in`); the merge reasoning is carried inline in
`Why they were merged` so the audit trail travels with the row.

Usage: 60_flat_sheet.py <outdir>
"""
import sys, json, csv, os, subprocess

OUT = sys.argv[1] if len(sys.argv) > 1 else '.'
HERE = os.path.dirname(os.path.abspath(__file__))
SP = '/tmp/claude-0/-home-user-Direct-Business/87d1aec0-e2cf-54b0-95a9-480203ce0e75/scratchpad'

# The workbook builder holds all the merge/verify logic; reuse it rather than
# duplicating, then flatten what it produced.
import openpyxl
XLSX = os.path.join(SP, 'TravelAgencies_MASTER_v4.xlsx')
if not os.path.exists(XLSX):
    subprocess.run([sys.executable, f'{HERE}/50_build_final.py', SP], check=True)

wb = openpyxl.load_workbook(XLSX, read_only=True)
comp = wb['Companies']
comp_hdr = [c.value for c in next(comp.iter_rows(min_row=1, max_row=1))]
people = wb['People, not companies']
people_hdr = [c.value for c in next(people.iter_rows(min_row=1, max_row=1))]
dupes = wb['Duplicates']
dupe_hdr = [c.value for c in next(dupes.iter_rows(min_row=1, max_row=1))]

# merge reasoning, keyed by the surviving row, so it can travel inline
why = {}
for r in dupes.iter_rows(min_row=2, values_only=True):
    d = dict(zip(dupe_hdr, r))
    k = str(d.get('Kept') or '')
    if k and k not in ('(none)', '(all kept separate)'):
        why[k] = f"{d.get('Decision','')} — {str(d.get('Why') or '')[:400]}"

COLS = list(comp_hdr) + ['Why they were merged']

rows = []
for r in comp.iter_rows(min_row=2, values_only=True):
    d = dict(zip(comp_hdr, r))
    d['Why they were merged'] = why.get(str(d.get('Row ID') or ''), '')
    rows.append(d)

# People go into the SAME table, so nothing lives on a tab the owner won't open.
pmap = {'Row ID': 'Row ID', 'Email 1 (deliverable)': 'Email 1 (deliverable)',
        'Email 2 (deliverable)': 'Email 2 (deliverable)', 'Mobile 1': 'Mobile 1',
        'Contact Names': 'Contact Names',
        'Where this record came from': 'Where this record came from'}
for r in people.iter_rows(min_row=2, values_only=True):
    d = dict(zip(people_hdr, r))
    row = {c: '' for c in COLS}
    for src, dst in pmap.items():
        if src in d: row[dst] = d.get(src) or ''
    row['Record Type'] = 'person — not a company'
    row['Verification Level'] = 'unverified'
    row['Where the name came from'] = ('Only identity is a personal email address, '
                                       'so this is a person, not a business')
    row['Priority Score'] = 0
    row['Completeness %'] = 0
    rows.append(row)

# Sort so the best prospects sit at the top and the people sit at the bottom.
def key(d):
    person = 1 if str(d.get('Record Type', '')).startswith('person') else 0
    return (person, -(d.get('Priority Score') or 0), -(d.get('Completeness %') or 0),
            str(d.get('Company Name (EN)') or d.get('Company Name (AR)') or ''))
rows.sort(key=key)

csv_path = os.path.join(OUT, 'TravelAgencies_ONE_TABLE.csv')
with open(csv_path, 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=COLS, extrasaction='ignore')
    w.writeheader()
    for d in rows:
        w.writerow({c: ('' if d.get(c) is None else d.get(c)) for c in COLS})

size = os.path.getsize(csv_path)
print(f'rows: {len(rows)} (companies + people in ONE table)')
print(f'columns: {len(COLS)}')
print(f'file: {csv_path}  ({size/1024/1024:.2f} MB)')
import collections
print('record types:', dict(collections.Counter(str(d.get('Record Type') or '') for d in rows)))
